#!/usr/bin/env python3
"""
Social Data Parser - Sidekick Social Audit
Version: 2.0
Parses CSV and Excel exports from Instagram, Facebook, and Google Business Profile.

Key Features:
- CSV and Excel (.xlsx) support
- Content-based deduplication (not just date+platform)
- Column header-based platform detection fallback
- Unique post_id for tracking
"""
import argparse
import json
import csv
import hashlib
import logging
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# Try importing openpyxl for Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class SocialDataParser:
    """Parses and normalizes social media export data."""

    # Column patterns that indicate a specific platform
    PLATFORM_COLUMN_HINTS = {
        'instagram': ['instagram', 'ig reach', 'ig impressions', 'reels'],
        'facebook': ['facebook', 'fb reach', 'fb impressions', 'page likes'],
        'google_business_profile': ['gbp', 'google business', 'local post', 'store code'],
    }

    def __init__(self, verbose=True, default_platform=None):
        self.verbose = verbose
        self.default_platform = default_platform
        self.stats = {
            'files_processed': 0,
            'files_failed': 0,
            'files_skipped': 0,
            'posts_parsed': 0,
            'posts_skipped': 0,
            'duplicates_removed': 0,
        }
        self.skipped_files = []  # Track why files were skipped

        if not EXCEL_SUPPORT:
            logger.warning("openpyxl not installed. Excel files will be skipped. Run: pip install openpyxl")

    def _generate_post_id(self, post):
        """Generate unique ID from date + platform + content hash."""
        content = post.get('caption', '') or post.get('post_type', '') or ''
        raw = f"{post['date']}_{post['platform']}_{content[:100]}"
        return hashlib.md5(raw.encode()).hexdigest()[:12]

    def _smart_deduplicate(self, posts):
        """Deduplicate by post_id (content-aware), not just date+platform."""
        seen = {}
        duplicates = 0

        for post in posts:
            post_id = self._generate_post_id(post)
            post['post_id'] = post_id

            if post_id in seen:
                # Keep the version with more data
                existing = seen[post_id]
                existing_fields = sum(1 for v in existing.values() if v)
                new_fields = sum(1 for v in post.values() if v)
                if new_fields > existing_fields:
                    seen[post_id] = post
                duplicates += 1
            else:
                seen[post_id] = post

        self.stats['duplicates_removed'] = duplicates
        return list(seen.values())

    def _detect_platform_from_columns(self, headers):
        """Detect platform from column names when filename doesn't indicate it."""
        headers_lower = ' '.join(str(h).lower() for h in headers if h)

        for platform, hints in self.PLATFORM_COLUMN_HINTS.items():
            for hint in hints:
                if hint in headers_lower:
                    return platform
        return None

    def _detect_platform(self, filename, headers=None):
        """Detect platform from filename, then fallback to column headers."""
        fn = filename.lower()

        # 1. Check filename first (most reliable)
        if 'gbp' in fn or 'google_business' in fn or 'googlebusiness' in fn or '_gbp-' in fn or '_gbp_' in fn:
            return 'google_business_profile'
        if 'instagram' in fn or '_ig_' in fn or '_ig-' in fn or 'ig_' in fn or '-ig-' in fn:
            return 'instagram'
        if 'facebook' in fn or '_fb_' in fn or '_fb-' in fn or 'fb_' in fn or '-fb-' in fn:
            return 'facebook'

        # 2. Fallback to column headers
        if headers:
            detected = self._detect_platform_from_columns(headers)
            if detected:
                return detected

        # 3. Use default platform if specified
        if self.default_platform:
            return self.default_platform

        return None

    def _normalize_post(self, row, filename, platform):
        """Normalize a single row into standard post format."""
        # Skip aggregate/summary rows
        if any(k in row for k in ['Store code', 'Business name']):
            if platform == 'google_business_profile':
                # These are valid GBP columns, don't skip the whole row
                pass
            else:
                return None

        # Normalize keys to lower case for searching
        row_map = {str(k).lower().strip(): k for k in row.keys() if k}

        # Date Finding Logic
        date = None
        date_keys = ['publish time', 'publish date', 'posted date', 'created', 'date', 'timestamp', 'post date']
        for key in date_keys:
            if key in row_map:
                raw_val = str(row[row_map[key]] or '')
                if 'lifetime' in raw_val.lower() or not raw_val.strip():
                    continue

                # Try multiple date formats
                for fmt in [
                    '%m/%d/%Y %H:%M', '%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S',
                    '%B %d, %Y', '%Y-%m-%d %H:%M', '%d/%m/%Y', '%Y/%m/%d'
                ]:
                    try:
                        date = datetime.strptime(raw_val.strip(), fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue
                if date:
                    break

        if not date:
            return None

        post = {
            'date': date,
            'platform': platform,
            'source_file': filename,
        }

        # Metric Mapping
        mapping = {
            'likes': ['likes', 'reactions', 'like count', 'total likes'],
            'comments': ['comments', 'comment count', 'total comments'],
            'shares': ['shares', 'share count', 'total shares'],
            'reach': ['reach', 'impressions', 'views', 'total reach'],
            'post_type': ['type', 'format', 'post type', 'media type'],
            'caption': ['caption', 'text', 'description', 'message', 'content'],
        }

        for field, lookup_keys in mapping.items():
            for key in lookup_keys:
                if key in row_map:
                    val = row[row_map[key]]
                    try:
                        if field in ['likes', 'comments', 'shares', 'reach']:
                            cleaned = str(val or '0').replace(',', '').replace('$', '').strip()
                            post[field] = int(float(cleaned or 0))
                        else:
                            post[field] = str(val or '').strip()
                    except (ValueError, TypeError):
                        post[field] = 0 if field in ['likes', 'comments', 'shares', 'reach'] else ''
                    break

        return post

    def _is_garbage_row(self, row):
        """Detects summary/aggregate rows often found in exports."""
        s = ' '.join(str(v).lower() for v in row.values() if v)
        garbage_indicators = ['number of', 'total count', 'interactions with', 'sum of', 'average']
        return any(x in s for x in garbage_indicators)

    def parse_csv(self, path):
        """Parse a CSV file into normalized posts."""
        posts = []
        try:
            with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
                sample = f.read(2048)
                f.seek(0)
                delim = ',' if sample.count(',') > sample.count(';') else ';'
                reader = csv.DictReader(f, delimiter=delim)
                headers = reader.fieldnames or []

                platform = self._detect_platform(path.name, headers)
                if not platform:
                    self.stats['files_skipped'] += 1
                    self.skipped_files.append({
                        'file': path.name,
                        'reason': 'Could not detect platform from filename or columns'
                    })
                    return []

                for row in reader:
                    if self._is_garbage_row(row):
                        continue
                    p = self._normalize_post(row, path.name, platform)
                    if p:
                        posts.append(p)
                        self.stats['posts_parsed'] += 1
                    else:
                        self.stats['posts_skipped'] += 1

            self.stats['files_processed'] += 1
        except Exception as e:
            self.stats['files_failed'] += 1
            logger.error(f"Failed CSV {path.name}: {e}")
        return posts

    def parse_excel(self, path):
        """Parse an Excel file into normalized posts."""
        posts = []
        if not EXCEL_SUPPORT:
            return []

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue

                headers = [str(h) if h else '' for h in rows[0]]
                platform = self._detect_platform(path.name, headers)

                if not platform:
                    # Try detecting from sheet name
                    sheet_lower = sheet_name.lower()
                    if 'instagram' in sheet_lower or 'ig' in sheet_lower:
                        platform = 'instagram'
                    elif 'facebook' in sheet_lower or 'fb' in sheet_lower:
                        platform = 'facebook'
                    elif 'gbp' in sheet_lower or 'google' in sheet_lower:
                        platform = 'google_business_profile'

                if not platform:
                    self.skipped_files.append({
                        'file': f"{path.name} / {sheet_name}",
                        'reason': 'Could not detect platform'
                    })
                    continue

                for row_data in rows[1:]:
                    row_dict = {}
                    for i, h in enumerate(headers):
                        if h and i < len(row_data):
                            row_dict[h] = row_data[i]

                    if self._is_garbage_row(row_dict):
                        continue

                    p = self._normalize_post(row_dict, path.name, platform)
                    if p:
                        posts.append(p)
                        self.stats['posts_parsed'] += 1
                    else:
                        self.stats['posts_skipped'] += 1

            if posts:
                self.stats['files_processed'] += 1
            else:
                self.stats['files_skipped'] += 1

        except Exception as e:
            self.stats['files_failed'] += 1
            logger.error(f"Failed Excel {path.name}: {e}")

        return posts

    def run(self, search_dir, output):
        """Run the full parsing pipeline."""
        posts = []
        path = Path(search_dir)

        if not path.exists():
            logger.error(f"Directory not found: {search_dir}")
            return

        # 0. First, enumerate ALL files in directory (nothing missed)
        all_files = list(path.rglob('*'))
        all_files = [f for f in all_files if f.is_file()]

        # Categorize files
        processable_extensions = {'.csv', '.xlsx', '.xls'}
        unprocessable_extensions = {'.json', '.md', '.txt', '.pdf', '.doc', '.docx', '.png', '.jpg', '.jpeg'}
        skip_patterns = {'.DS_Store', '__pycache__', '.git'}

        file_manifest = {
            'total_found': len(all_files),
            'processable': [],
            'unprocessable': [],
            'skipped': [],
        }

        for f in all_files:
            if any(skip in str(f) for skip in skip_patterns):
                file_manifest['skipped'].append((f.name, 'system file'))
                continue
            ext = f.suffix.lower()
            if ext in processable_extensions:
                file_manifest['processable'].append(f)
            elif ext in unprocessable_extensions:
                file_manifest['unprocessable'].append((f.name, f'not social data ({ext})'))
            elif ext:
                file_manifest['skipped'].append((f.name, f'unknown ({ext})'))

        # Log complete file inventory
        logger.info(f"📁 FILE INVENTORY: {file_manifest['total_found']} total files found")
        logger.info(f"   ✓ Processable: {len(file_manifest['processable'])} files")
        if file_manifest['unprocessable']:
            logger.info(f"   ⚠️ Unprocessable: {len(file_manifest['unprocessable'])} files")
        if file_manifest['skipped']:
            logger.info(f"   ⏭️ Skipped: {len(file_manifest['skipped'])} files")

        # Store for output
        self.file_manifest = file_manifest

        # 1. Parse CSVs
        csv_files = [f for f in file_manifest['processable'] if f.suffix.lower() == '.csv']
        for f in csv_files:
            posts.extend(self.parse_csv(f))

        # 2. Parse Excels
        if EXCEL_SUPPORT:
            xlsx_files = [f for f in file_manifest['processable'] if f.suffix.lower() in {'.xlsx', '.xls'}]
            for f in xlsx_files:
                posts.extend(self.parse_excel(f))

        # 3. Deduplicate (content-aware)
        unique = self._smart_deduplicate(posts)
        unique.sort(key=lambda x: x['date'])

        # 4. Save output
        out = Path(output)
        out.parent.mkdir(parents=True, exist_ok=True)

        result = {
            'posts': unique,
            'stats': self.stats,
            'skipped_files': self.skipped_files,
            'file_manifest': {
                'total_found': self.file_manifest['total_found'],
                'processable_count': len(self.file_manifest['processable']),
                'processable_files': [str(f.name) for f in self.file_manifest['processable']],
                'unprocessable': self.file_manifest['unprocessable'],
                'skipped': self.file_manifest['skipped'],
            },
            'generated_at': datetime.now().isoformat(),
        }

        with open(out, 'w') as f:
            json.dump(result, f, indent=2, default=str)

        # 5. Summary logging
        logger.info("=" * 60)
        logger.info("✅ PARSING COMPLETE")
        logger.info(f"   📊 Parsed {len(unique)} unique posts from {self.stats['files_processed']} files")
        if self.stats['duplicates_removed']:
            logger.info(f"   🔄 Removed {self.stats['duplicates_removed']} duplicates")
        if self.stats['files_skipped']:
            logger.warning(f"   ⏭️ Skipped {self.stats['files_skipped']} files (unknown platform)")
        if self.stats['files_failed']:
            logger.error(f"   ❌ Failed {self.stats['files_failed']} files")

        # File manifest summary
        if self.file_manifest['unprocessable']:
            logger.warning(f"   ⚠️ REVIEW MANUALLY: {len(self.file_manifest['unprocessable'])} files not processed:")
            for fname, reason in self.file_manifest['unprocessable'][:5]:
                logger.warning(f"      - {fname}: {reason}")
            if len(self.file_manifest['unprocessable']) > 5:
                logger.warning(f"      ... and {len(self.file_manifest['unprocessable']) - 5} more")

        if self.skipped_files and self.verbose:
            logger.info("Platform detection failures:")
            for skip in self.skipped_files:
                logger.info(f"  - {skip['file']}: {skip['reason']}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Parse social media export files (CSV, Excel) into normalized JSON'
    )
    parser.add_argument('--search-dir', required=True, help='Directory containing export files')
    parser.add_argument('--output', required=True, help='Output JSON file path')
    parser.add_argument(
        '--platform',
        choices=['instagram', 'facebook', 'google_business_profile'],
        help='Default platform if auto-detection fails'
    )
    parser.add_argument('--quiet', action='store_true', help='Reduce logging output')

    args = parser.parse_args()

    parser_instance = SocialDataParser(
        verbose=not args.quiet,
        default_platform=args.platform
    )
    parser_instance.run(args.search_dir, args.output)
